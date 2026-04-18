import openpyxl
import random
from django.contrib.auth.models import User, Group , Permission
from django.contrib.contenttypes.models import ContentType
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Sum
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm

from .models import Transaction, Entreprise, CompteComptable, EcritureComptable, Client, Facture
# ==========================================
# 1. AUTHENTIFICATION & DECONNEXION
# ==========================================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Identifiant ou mot de passe incorrect.")
            
    return render(request, 'authentification.html')

def logout_view(request):
    logout(request)
    return redirect('login') 

# ==========================================
# 2. DASHBOARD
# ==========================================
# ==========================================
# 2. DASHBOARD (VRAIES DONNÉES)
# ==========================================
@login_required(login_url='login')
def dashboard(request):
    entreprise = Entreprise.objects.filter(gerant=request.user).first()
    
    # Si l'utilisateur vient de se créer et n'a pas encore de données
    if not entreprise:
        return render(request, 'dashboard.html', {
            'ca_total': 0, 'ca_encaisse': 0, 'resultat_net': 0, 'dernieres_factures': []
        })

    # 1. Données de Facturation
    factures = Facture.objects.filter(entreprise=entreprise)
    ca_total = factures.aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0
    ca_encaisse = factures.filter(statut='PAYE').aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0
    dernieres_factures = factures.order_by('-date_creation')[:5] # Les 5 plus récentes

    # 2. Données Comptables (Résultat Net : Produits - Charges)
    ecritures = EcritureComptable.objects.filter(entreprise=entreprise)
    total_produits = ecritures.filter(compte__classe=7).aggregate(Sum('credit'))['credit__sum'] or 0
    total_charges = ecritures.filter(compte__classe=6).aggregate(Sum('debit'))['debit__sum'] or 0
    resultat_net = total_produits - total_charges

    context = {
        'ca_total': float(ca_total),
        'ca_encaisse': float(ca_encaisse),
        'resultat_net': float(resultat_net),
        'dernieres_factures': dernieres_factures,
    }
    return render(request, 'dashboard.html', context)

# ==========================================
# 3. MODULE COMPTABILITÉ
# ==========================================
@login_required(login_url='login')
def comptabilite_view(request):
    entreprise, created = Entreprise.objects.get_or_create(
        gerant=request.user,
        defaults={'nom': f"Entreprise de {request.user.username}"}
    )

    ecritures = EcritureComptable.objects.filter(entreprise=entreprise).order_by('-date')
    comptes = CompteComptable.objects.all().order_by('numero')

    total_produits = ecritures.filter(compte__classe=7).aggregate(Sum('credit'))['credit__sum'] or 0
    total_charges = ecritures.filter(compte__classe=6).aggregate(Sum('debit'))['debit__sum'] or 0
    resultat_net = total_produits - total_charges

    total_actif = (ecritures.filter(compte__classe__in=[2,3,4,5]).aggregate(Sum('debit'))['debit__sum'] or 0) - \
                  (ecritures.filter(compte__classe__in=[2,3,4,5]).aggregate(Sum('credit'))['credit__sum'] or 0)

    passif_c = ecritures.filter(compte__classe=1).aggregate(Sum('credit'))['credit__sum'] or 0
    passif_d = ecritures.filter(compte__classe=1).aggregate(Sum('debit'))['debit__sum'] or 0
    total_passif = (passif_c - passif_d) + resultat_net

    context = {
        'ecritures': ecritures,
        'comptes': comptes,
        'resultat_net': resultat_net,
        'total_actif': total_actif,
        'total_passif': total_passif,
    }
    return render(request, 'comptabilite.html', context)

@login_required(login_url='login')
def ajouter_ecriture(request):
    if request.method == 'POST':
        entreprise = Entreprise.objects.filter(gerant=request.user).first()
        if entreprise:
            EcritureComptable.objects.create(
                entreprise=entreprise,
                date=request.POST.get('date'),
                libelle=request.POST.get('libelle'),
                compte_id=request.POST.get('compte'),
                debit=float(request.POST.get('debit') or 0),
                credit=float(request.POST.get('credit') or 0)
            )
            messages.success(request, "L'écriture a bien été enregistrée.")
    return redirect('comptabilite')

@login_required(login_url='login')
def supprimer_ecriture(request, id):
    if not request.user.is_superuser:
        messages.error(request, "Tu n'as pas le droit de suppression.")
        return redirect('comptabilite')
    
    get_object_or_404(EcritureComptable, id=id).delete()
    messages.success(request, "L'écriture a été supprimée.")
    return redirect('comptabilite')

# ==========================================
# 4. MODULE FACTURATION
# ==========================================

@login_required(login_url='login')
def facturation_view(request):
    entreprise = Entreprise.objects.filter(gerant=request.user).first()
    if not entreprise:
        return redirect('dashboard')

    factures = Facture.objects.filter(entreprise=entreprise).order_by('-date_creation')
    
    # NOUVEAU : On récupère la liste des clients pour le menu déroulant
    clients = Client.objects.filter(entreprise=entreprise).order_by('nom_client')
    
    # Calcul des KPI automatiques
    total_facture = factures.aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0
    total_paye = factures.filter(statut='PAYE').aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0
    total_attente = factures.exclude(statut='PAYE').aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0

    context = {
        'factures': factures,
        'clients': clients, # Ajouté ici
        'total_facture': total_facture,
        'total_paye': total_paye,
        'total_attente': total_attente,
    }
    return render(request, 'facturation.html', context)

@login_required(login_url='login')
def ajouter_client(request):
    if request.method == 'POST':
        entreprise = Entreprise.objects.filter(gerant=request.user).first()
        
        # Génération du numéro de client automatique (ex: CLI-0001)
        count = Client.objects.filter(entreprise=entreprise).count() + 1
        numero_genere = f"CLI-{count:04d}"

        Client.objects.create(
            entreprise=entreprise,
            numero_client=numero_genere,
            nom_client=request.POST.get('nom_client'),
            email_client=request.POST.get('email_client'),
            telephone=request.POST.get('telephone'),
            adresse=request.POST.get('adresse'),
            description=request.POST.get('description')
        )
        messages.success(request, f"Le client {numero_genere} a bien été ajouté.")
    return redirect('facturation')

@login_required(login_url='login')
def supprimer_facture(request, id):
    facture = get_object_or_404(Facture, id=id)
    facture.delete()
    messages.success(request, "La facture a été supprimée avec succès.")
    return redirect('facturation')

@login_required(login_url='login')
def detail_facture(request, id):
    # Pour l'instant, on redirige avec un message info (à développer plus tard)
    messages.info(request, "L'interface de visualisation arrive bientôt !")
    return redirect('facturation')

@login_required(login_url='login')
def creer_facture(request):
    if request.method == 'POST':
        entreprise = Entreprise.objects.filter(gerant=request.user).first()
        client_id = request.POST.get('client')
        date_echeance = request.POST.get('date_echeance')
        montant_ttc = request.POST.get('montant_ttc')
        statut = request.POST.get('statut')

        if entreprise and client_id:
            client = get_object_or_404(Client, id=client_id)
            
            # Génération automatique du numéro de facture (ex: FAC-2026-0001)
            annee = timezone.now().year
            count = Facture.objects.filter(entreprise=entreprise).count() + 1
            numero_genere = f"FAC-{annee}-{count:04d}"

            Facture.objects.create(
                entreprise=entreprise,
                numero_facture=numero_genere,
                client=client,
                date_echeance=date_echeance,
                montant_ttc=montant_ttc,
                statut=statut
            )
            messages.success(request, f"La facture {numero_genere} a bien été créée.")
    
    return redirect('facturation')

# ==========================================
# 5. MODULE FACTURATION (GÉNÉRATION PDF)
@login_required(login_url='login')
def pdf_facture(request, id):
    # Même logique que pour le justificatif, version facturation
    facture = get_object_or_404(Facture, id=id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Facture_{facture.numero_facture}.pdf"'
    
    p = canvas.Canvas(response, pagesize=A4)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(2 * cm, 27 * cm, f"FACTURE {facture.numero_facture}")
    p.setFont("Helvetica", 12)
    p.drawString(2 * cm, 25 * cm, f"Client : {facture.client.nom_client}")
    p.drawString(2 * cm, 24 * cm, f"Montant : {facture.montant_ttc} EUROS")
    p.drawString(2 * cm, 23 * cm, f"Statut : {facture.statut}")
    p.showPage()
    p.save()
    return response

# ==========================================
# 
# 6.  EXPORTS EXCEL
# ==========================================
@login_required(login_url='login')
def export_excel_compta(request):
    entreprise = Entreprise.objects.filter(gerant=request.user).first()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Date', 'N° Compte', 'Libelle', 'Debit', 'Credit'])
    for e in EcritureComptable.objects.filter(entreprise=entreprise).order_by('-date'):
        ws.append([e.date.strftime("%d/%m/%Y"), e.compte.numero, e.libelle, e.debit, e.credit])
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=journal_{request.user.username}.xlsx'
    wb.save(response)
    return response
# ==========================================
# 7. IMPORTS  EXCEL
@login_required(login_url='login')
def import_excel_compta(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file or not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Fichier invalide.")
            return redirect('comptabilite')
        try:
            entreprise = Entreprise.objects.filter(gerant=request.user).first()
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                compte = CompteComptable.objects.filter(numero=str(row[1]).strip()).first()
                if compte:
                    EcritureComptable.objects.create(
                        entreprise=entreprise,
                        date=row[0] if isinstance(row[0], datetime) else datetime.today().date(),
                        libelle=str(row[2]), compte=compte, debit=float(row[3] or 0), credit=float(row[4] or 0)
                    )
                    count += 1
            messages.success(request, f"{count} lignes importées.")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
    return redirect('comptabilite')

# ==========================================
# 8. GÉNÉRATION PDF & SYNC BANQUE
# ==========================================
@login_required(login_url='login')
def generer_facture_pdf(request, ecriture_id):
    ecriture = get_object_or_404(EcritureComptable, id=ecriture_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Justificatif_{ecriture.id}.pdf"'
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    p.setFont("Helvetica-Bold", 16)
    p.drawString(2 * cm, height - 2 * cm, "JUSTIFICATIF D'ÉCRITURES COMPTABLES")
    p.setFont("Helvetica", 11)
    p.drawString(2 * cm, height - 3.5 * cm, f"Entreprise : {ecriture.entreprise.nom}")
    p.drawString(2 * cm, height - 4 * cm, f"Date : {ecriture.date.strftime('%d/%m/%Y')}")
    p.line(2 * cm, height - 4.5 * cm, width - 2 * cm, height - 4.5 * cm)
    p.drawString(2 * cm, height - 6 * cm, f"Libellé : {ecriture.libelle}")
    p.drawString(2 * cm, height - 6.5 * cm, f"Compte : {ecriture.compte.numero} - {ecriture.compte.libelle}")
    montant = ecriture.debit if ecriture.debit > 0 else ecriture.credit
    p.setFont("Helvetica-Bold", 12)
    p.drawString(2 * cm, height - 8 * cm, f"Montant Total : {montant} €")
    p.showPage()
    p.save()
    return response



# sync banque (génération de transactions aléatoires pour simuler une API bancaire)
@login_required(login_url='login')
def sync_bank_api(request):
    entreprise = Entreprise.objects.filter(gerant=request.user).first()
    if not entreprise: return redirect('dashboard')
    lib_out = ["Abonnement SaaS", "Facture EDF", "Fournisseur Matériel"]
    lib_in = ["Règlement Client Dupont", "Virement Client A"]
    for _ in range(3):
        t_type = random.choice(['IN', 'OUT'])
        Transaction.objects.create(
            entreprise=entreprise, date=timezone.now().date(),
            libelle=f"[SYNC] {random.choice(lib_in if t_type == 'IN' else lib_out)}",
            type=t_type, montant=round(random.uniform(100, 2000), 2)
        )
    messages.success(request, "Synchronisation bancaire réussie.")
    return redirect('dashboard')




# ==========================================
# 9. MODULE REPORTING / ANALYSE
# ==========================================
@login_required(login_url='login')
def reporting_view(request):
    entreprise = Entreprise.objects.filter(gerant=request.user).first()
    if not entreprise:
        messages.warning(request, "Veuillez d'abord configurer votre entreprise.")
        return redirect('dashboard')

    # --- a. Données Facturation ---
    factures = Facture.objects.filter(entreprise=entreprise)
    ca_paye = factures.filter(statut='PAYE').aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0
    ca_attente = factures.filter(statut='ATTENTE').aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0
    ca_envoye = factures.filter(statut='ENVOYE').aggregate(Sum('montant_ttc'))['montant_ttc__sum'] or 0

    # --- b. Données Comptables (Charges vs Produits) ---
    ecritures = EcritureComptable.objects.filter(entreprise=entreprise)
    total_charges = ecritures.filter(compte__classe=6).aggregate(Sum('debit'))['debit__sum'] or 0
    total_produits = ecritures.filter(compte__classe=7).aggregate(Sum('credit'))['credit__sum'] or 0

    # --- c. Données de Trésorerie (Transactions IN / OUT) ---
    transactions = Transaction.objects.filter(entreprise=entreprise)
    total_in = transactions.filter(type='IN').aggregate(Sum('montant'))['montant__sum'] or 0
    total_out = transactions.filter(type='OUT').aggregate(Sum('montant'))['montant__sum'] or 0

    # On passe les données en format "float" pour que les graphiques Javascript puissent les lire
    context = {
        'ca_paye': float(ca_paye),
        'ca_attente': float(ca_attente),
        'ca_envoye': float(ca_envoye),
        'total_charges': float(total_charges),
        'total_produits': float(total_produits),
        'total_in': float(total_in),
        'total_out': float(total_out),
    }
    
    return render(request, 'reporting.html', context)


# ==========================================
# 10. MODULE ADMINISTRATION (PERSONNALISÉ)
@login_required(login_url='login')
def administration_view(request):
    if not request.user.is_superuser:
        messages.error(request, "Accès refusé. Vous n'avez pas les droits d'administration.")
        return redirect('dashboard')

    utilisateurs = User.objects.all().order_by('-date_joined')
    groupes = Group.objects.all()

    return render(request, 'administration.html', {
        'utilisateurs': utilisateurs,
        'groupes': groupes,
    })

@login_required(login_url='login')
def ajouter_utilisateur(request):
    if request.method == 'POST' and request.user.is_superuser:
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role') # Peut être 'admin' ou l'ID d'un groupe

        if User.objects.filter(username=username).exists():
            messages.error(request, f"L'identifiant '{username}' existe déjà.")
            return redirect('administration_custom')

        # Création de l'utilisateur avec mot de passe sécurisé (haché automatiquement)
        user = User.objects.create_user(username=username, password=password)

        # Attribution du rôle / groupe
        if role == 'admin':
            user.is_superuser = True
            user.is_staff = True
            user.save()
        elif role:
            try:
                groupe = Group.objects.get(id=role)
                user.groups.add(groupe)
            except Group.DoesNotExist:
                pass

        messages.success(request, f"L'utilisateur {username} a été créé avec succès.")
    return redirect('administration_custom')

@login_required(login_url='login')
def modifier_utilisateur(request, user_id):
    if request.method == 'POST' and request.user.is_superuser:
        user = get_object_or_404(User, id=user_id)
        nouveau_username = request.POST.get('username')
        nouveau_password = request.POST.get('password')
        role = request.POST.get('role')

        # Vérifier que le nouveau pseudo n'est pas déjà pris par quelqu'un d'autre
        if User.objects.filter(username=nouveau_username).exclude(id=user_id).exists():
            messages.error(request, "Cet identifiant est déjà utilisé.")
            return redirect('administration_custom')

        user.username = nouveau_username
        
        # Si le champ mot de passe n'est pas vide, on le change (hachage auto)
        if nouveau_password:
            user.set_password(nouveau_password)

        # Nettoyage et mise à jour des rôles
        user.is_superuser = False
        user.is_staff = False
        user.groups.clear()

        if role == 'admin':
            user.is_superuser = True
            user.is_staff = True
        elif role:
            try:
                groupe = Group.objects.get(id=role)
                user.groups.add(groupe)
            except Group.DoesNotExist:
                pass
        
        user.save()
        messages.success(request, f"Le profil de {user.username} a été mis à jour.")
    return redirect('administration_custom')

@login_required(login_url='login')
def supprimer_utilisateur(request, user_id):
    if not request.user.is_superuser:
        return redirect('dashboard')
        
    user = get_object_or_404(User, id=user_id)
    if user == request.user:
        messages.error(request, "Vous ne pouvez pas supprimer votre propre compte !")
    else:
        user.delete()
        messages.success(request, "Utilisateur supprimé de la base de données.")
        
    return redirect('administration_custom')

# ==========================================
# 11. ADMINISTRATION : USERS & GROUPES
# ==========================================

@login_required(login_url='login')
def administration_view(request):
    if not request.user.is_superuser:
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')

    utilisateurs = User.objects.all().order_by('-date_joined')
    groupes = Group.objects.all()
    # On récupère les permissions liées à ton application 'authentification'
    # pour ne pas afficher les centaines de permissions par défaut de Django
    permissions = Permission.objects.filter(content_type__app_label='authentification')

    return render(request, 'administration.html', {
        'utilisateurs': utilisateurs,
        'groupes': groupes,
        'permissions': permissions,
    })

@login_required(login_url='login')
def ajouter_groupe(request):
    if request.method == 'POST' and request.user.is_superuser:
        nom_groupe = request.POST.get('name')
        perms_ids = request.POST.getlist('permissions')
        
        if Group.objects.filter(name=nom_groupe).exists():
            messages.error(request, f"Le groupe '{nom_groupe}' existe déjà.")
        else:
            nouveau_groupe = Group.objects.create(name=nom_groupe)
            # On ajoute les permissions sélectionnées
            if perms_ids:
                nouveau_groupe.permissions.set(perms_ids)
            messages.success(request, f"Groupe '{nom_groupe}' créé avec succès.")
            
    return redirect('administration_custom')

@login_required(login_url='login')
def supprimer_groupe(request, group_id):
    if request.user.is_superuser:
        groupe = get_object_or_404(Group, id=group_id)
        groupe.delete()
        messages.success(request, "Groupe supprimé.")
    return redirect('administration_custom')